# -*- coding: utf-8 -*-
"""
Created on Wed Jan 20 00:36:29 2021

@author: Jainan
"""

import openpyxl, pprint, json

wb = openpyxl.load_workbook('Final_Data.xlsx')

sheet1 = wb['Qsn_Tag']
sheet2 = wb['Responses']

excel_data = {}
excel_data["intents"] = []

tag={}

for row in range(3, sheet1.max_row  + 1):
    Col_C  = sheet1['C' + str(row)].value
    Col_I = sheet1['I' + str(row)].value
    if str(Col_I) != "None":
        if Col_I not in tag:
            tag.update({Col_I: [Col_C]})
        else:
            tag[Col_I].append(Col_C)

dict2={}

for row in range(3, sheet2.max_row +1):
    Col_D = sheet2['D' + str(row)].value
    Col_G = sheet2['G' + str(row)].value
    if str(Col_D) != "None":
        dict2.update({Col_D: Col_G})

# print([ x for x in list(tag.keys()) if x not in list(dict2.keys())])

# print([ x for x in list(dict2.keys()) if x not in list(tag.keys())])

for record in tag.keys():
        excel_data["intents"].append({"tag": record, "patterns": tag[record], "response": [dict2[record]]})

# Open a new text file and write the contents of excel_data to it.

with open('DATA.json', 'w') as resultFile:
    json.dump(excel_data, resultFile)
